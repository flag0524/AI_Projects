"""
샘플 product_data.xlsx 생성 스크립트
실제 데이터를 입력하기 전에 엑셀 파일의 구조를 확인용으로 사용하세요.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def create_sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품데이터"

    # ─── 헤더 스타일 ───
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color="CCCCCC"),
        right=Side(style='thin', color="CCCCCC"),
        top=Side(style='thin', color="CCCCCC"),
        bottom=Side(style='thin', color="CCCCCC")
    )

    headers = ["품번", "품명", "카테고리", "최초판매가", "색상", "시즌", "코디그룹"]
    col_widths = [15, 25, 12, 15, 12, 10, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30

    # ─── 샘플 데이터 ───
    sample_data = [
        ["BLD-2401", "린넨 오버핏 재킷",    "아우터", 89000,  "아이보리",  "2024SS", "C001"],
        ["BLD-2402", "와이드 린넨 팬츠",     "하의",   59000,  "베이지",    "2024SS", "C001,C003"],
        ["BLD-2403", "스트라이프 블라우스",   "상의",   49000,  "화이트",    "2024SS", "C002"],
        ["BLD-2404", "A라인 미디 스커트",     "하의",   65000,  "네이비",    "2024SS", "C002"],
        ["BLD-2405", "플로럴 쉬폰 원피스",   "원피스", 95000,  "플라워 멀티", "2024SS", "C003"],
        ["BLD-2406", "크롭 트위드 재킷",     "아우터", 129000, "블랙",      "2024FW", "C004"],
        ["BLD-2407", "슬림 슬랙스",          "하의",   69000,  "차콜",      "2024FW", "C004"],
        ["BLD-2408", "캐시미어 블렌드 니트", "상의",   79000,  "카멜",      "2024FW", "C004,C005"],
        ["BLD-2409", "롱 울 코트",           "아우터", 198000, "카멜",      "2024FW", "C005"],
        ["BLD-2410", "벨벳 미디 스커트",     "하의",   75000,  "버건디",    "2024FW", "C005"],
    ]

    # 데이터 행 스타일
    data_fill_even = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    data_align_center = Alignment(horizontal="center", vertical="center")
    price_align = Alignment(horizontal="right", vertical="center")

    for row_idx, row_data in enumerate(sample_data, start=2):
        fill = data_fill_even if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = price_align if col_idx == 4 else data_align_center
            if fill:
                cell.fill = fill
            if col_idx == 4:
                cell.number_format = '#,##0"원"'
        ws.row_dimensions[row_idx].height = 22

    # 주석 시트 추가
    ws_guide = wb.create_sheet("작성 가이드")
    guide_data = [
        ["컬럼명", "형식", "설명", "예시"],
        ["품번", "텍스트", "파일명과 반드시 동일하게 입력 (대소문자 구분)", "BLD-2401"],
        ["품명", "텍스트", "상품 정식 명칭", "린넨 오버핏 재킷"],
        ["카테고리", "텍스트", "아우터 / 상의 / 하의 / 원피스 중 하나", "아우터"],
        ["최초판매가", "숫자", "원 단위 숫자만 입력 (콤마, '원' 문자 제외)", "89000"],
        ["색상", "텍스트", "대표 색상명", "아이보리"],
        ["시즌", "텍스트", "시즌 코드", "2024SS"],
        ["코디그룹", "텍스트", "코디 폴더명. 복수는 쉼표로 구분", "C001 또는 C001,C003"],
    ]
    for row in guide_data:
        ws_guide.append(row)

    output_path = Path(__file__).parent.parent / "input" / "product_data.xlsx"
    wb.save(output_path)
    print(f"✅ 샘플 엑셀 파일 생성 완료: {output_path}")
    print(f"   → 상품 데이터를 직접 입력하거나, 기존 엑셀에서 복사하여 사용하세요.")

if __name__ == "__main__":
    create_sample_excel()
